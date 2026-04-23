"""
CountryZoning TXT Generator

HOW THIS FILE FITS INTO THE BIGGER PICTURE
-------------------------------------------
After create_table.py builds the Excel workbook, this script reads the
CountryZoning tab from that Excel file and produces a compact plain-text summary.

The TXT file lists each **CountryZoning** rate / zone label (left column) followed by
all **ISO codes** for that group (right column).  The left column is **not** read from
``addition/dhl_country_codes.txt`` — that file only maps **country names** to codes.
Short zone names like ``WW_EXP_IMP_ZONE_3`` come from the CountryZoning sheet (same
normalization as MainCosts: ``…_Zone_N`` not ``…_Zone_Zone N``).

  Express Worldwide  DE, FR, IT, ES, NL, BE
  WW_EXP_IMP_ZONE_3  DE, FR, …

When ``extracted_json_path`` is set and the JSON contains ``DemandSurchargeCountries``,
additional lines are appended after a blank line.  If a zone appears in both the
origin and destination country tables with the **same** country list, a single
``DemandSurcharge_<ZoneToken>`` line is written (from the origin block; the duplicate
in the destination block is skipped).  If the lists differ, that zone uses
``DemandSurcharge_Origin_*`` in the origin section and ``DemandSurcharge_Destination_*``
in the destination section.  Country names are resolved to 2-letter codes via
``addition/dhl_country_codes.txt`` where possible; unresolved segments are left as
in the source.

When the JSON contains ``GoGreenPlusCost``, a further blank line and GoGreen block
lines are appended: ``GoGreenOrigin_1  ES, IT, ...`` (and Destination / Origin_Destination
variants), matching the placeholders written to the GoGreenPlusCost Excel tab.

This file is useful for quickly checking which countries are covered by each rate
without opening the full Excel workbook.

Output is saved to the same folder as the Excel file (or to a custom path).
"""

import json
from pathlib import Path           # cross-platform file path handling
from collections import defaultdict  # used to build a dict of lists (rate name -> [countries])

from expand_additional_zoning import _normalize_zone_label
from transform_other_tabs import (
    _load_country_codes,
    _gogreen_country_list_to_codes,
    build_gogreen_block_txt_lines,
    demand_surcharge_global_label,
    demand_surcharge_origin_label,
    demand_surcharge_destination_label,
    parse_demand_surcharge_zone_country_maps,
    zone_uses_global_demand_surcharge_label,
)


def _append_demand_surcharge_countries_lines(lines, demand_surcharge_countries, name_to_code):
    """
    Append DemandSurchargeCountries block: one line per zone.

    Uses ``DemandSurcharge_<Token>`` when origin and destination country lists match
    for that zone; otherwise ``DemandSurcharge_Origin_*`` / ``DemandSurcharge_Destination_*``.
    """
    if not demand_surcharge_countries:
        return
    lines.append("")
    origin_map, dest_map = parse_demand_surcharge_zone_country_maps(
        demand_surcharge_countries
    )
    # Global zone lines: emit at first encounter (origin or destination), skip duplicate.
    seen_global = set()
    section = "origin"
    for row in demand_surcharge_countries:
        if not isinstance(row, dict):
            continue
        # Azure / extraction often uses "Origin/Destination" (slash); older code used underscore.
        od = (
            row.get("Origin_Destination")
            or row.get("origin_destination")
            or row.get("Origin/Destination")
            or row.get("origin/destination")
        )
        if od and str(od).strip():
            t = str(od).strip().lower()
            if "origin" in t and "territor" in t:
                section = "origin"
                continue
            if "destination" in t and "territor" in t:
                section = "destination"
                continue
        zn = row.get("ZoneName") or row.get("zoneName")
        if zn is None or not str(zn).strip():
            continue
        zone_name = str(zn).strip()
        raw = row.get("Countries") or row.get("countries") or ""
        if not isinstance(raw, str):
            raw = str(raw)
        use_global = zone_uses_global_demand_surcharge_label(
            zone_name, origin_map, dest_map
        )
        if use_global:
            if zone_name in seen_global:
                continue
            seen_global.add(zone_name)
            prefix = demand_surcharge_global_label(zone_name)
        else:
            if section == "destination":
                prefix = demand_surcharge_destination_label(zone_name)
            else:
                prefix = demand_surcharge_origin_label(zone_name)
        codes = _gogreen_country_list_to_codes(raw, name_to_code) if raw.strip() else ""
        lines.append(f"{prefix}  {codes}")


def _append_gogreen_block_lines(lines, go_green_array, name_to_code):
    """Append GoGreen placeholder block definitions (codes) from GoGreenPlusCost JSON."""
    if not go_green_array:
        return
    gg_lines = build_gogreen_block_txt_lines(go_green_array, name_to_code)
    if not gg_lines:
        return
    lines.append("")
    lines.extend(gg_lines)


def create_country_region_txt(
    excel_path: str = "output/DHL_Rate_Cards.xlsx",
    sheet_name: str = "CountryZoning",
    output_path: str | None = None,
    extracted_json_path: str | None = None,
) -> str:
    """
    Read the CountryZoning tab from an Excel workbook, group countries by rate name,
    and write a plain-text summary file.

    HOW IT WORKS:
      1. Open the Excel file and find the CountryZoning sheet.
      2. Read the header row to find which columns are "RateName" and "Country Code".
      3. Loop through all data rows, grouping country codes under their rate name.
         If a row's RateName cell is blank, the previous row's rate name is reused
         (forward-fill), because the Excel sheet may have merged cells for the rate name.
      4. Write one line per rate name: "RateName  code1, code2, code3, ..."
      5. Return the path of the created TXT file.

    Parameters:
      excel_path   – path to the Excel workbook to read (default: output/DHL_Rate_Cards.xlsx)
      sheet_name   – name of the sheet to read (default: "CountryZoning")
      output_path  – where to save the TXT file; if None, saves next to the Excel file
      extracted_json_path – optional path to extracted_data.json; if present and readable,
                            DemandSurchargeCountries entries are appended after CountryZoning lines.

    Returns the path of the created TXT file as a string.
    """
    # openpyxl is the library used to read Excel files; it's imported here (not at the top)
    # so that the rest of the project still works even if openpyxl isn't installed
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required. Install with: pip install openpyxl")

    excel_path = Path(excel_path)
    print(f"[*] TXT Debug: excel_path={excel_path}")
    print(f"[*] TXT Debug: sheet_name={sheet_name}")

    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    # Determine where to save the TXT file
    if output_path is None:
        # Default: save in the same folder as the Excel file
        output_dir = excel_path.parent
        output_path = output_dir / "CountryZoning_by_RateName.txt"
    else:
        output_path = Path(output_path)

    # Open the Excel file in read-only mode (faster; we only need to read, not write)
    # data_only=True means we get the cell values, not the formulas
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    lines = []

    # Check that the CountryZoning sheet exists in this workbook.
    # Some rate cards don't have country zoning data, so the sheet may be absent.
    if sheet_name not in wb.sheetnames:
        wb.close()
        print(
            f"[WARN] Sheet '{sheet_name}' not found in {excel_path} "
            f"(no CountryZoning tab). TXT will contain DemandSurchargeCountries only if JSON provides them."
        )
        rows = []
    else:
        print(f"[*] TXT Debug: workbook sheets={wb.sheetnames}")

        # Read all rows from the CountryZoning sheet into memory at once
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))   # each row is a tuple of cell values
        wb.close()
        print(f"[*] TXT Debug: total rows read (including header)={len(rows)}")

    if not rows:
        print("[WARN] TXT Debug: no CountryZoning rows to read (empty or missing sheet)")
    else:
        # -----------------------------------------------------------------------
        # Find the column positions for "RateName" and "Country Code".
        # We do this by reading the first row (the header row) and searching for
        # those exact column names.  This way the code still works even if the
        # columns are in a different order.
        # -----------------------------------------------------------------------
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        rate_name_col = None
        country_col = None
        for i, h in enumerate(headers):
            if h == "RateName":
                rate_name_col = i
            if h == "Country Code":
                country_col = i

        print(f"[*] TXT Debug: headers={headers}")
        print(f"[*] TXT Debug: RateName col index={rate_name_col}, Country Code col index={country_col}")

        if rate_name_col is None or country_col is None:
            print(
                "[WARN] TXT Debug: CountryZoning missing RateName or Country Code column; "
                "skipping Excel block"
            )
        else:
            # -----------------------------------------------------------------------
            # Loop through all data rows (skipping the header row) and group country
            # codes by rate name.
            #
            # FORWARD-FILL LOGIC:
            # In the Excel sheet, the RateName column may have blank cells for rows
            # that belong to the same rate as the row above (because the cells are
            # visually merged in Excel).  When we read the sheet, merged cells appear
            # as blank after the first row.  We handle this by remembering the last
            # non-blank rate name and reusing it for blank cells.
            #
            # Example:
            #   Row 2: RateName="Express Worldwide", Country Code="DE"
            #   Row 3: RateName="",                  Country Code="FR"   <- blank; reuse "Express Worldwide"
            #   Row 4: RateName="",                  Country Code="IT"   <- blank; reuse "Express Worldwide"
            #   Row 5: RateName="Economy Select",    Country Code="GB"   <- new rate name
            # -----------------------------------------------------------------------
            by_rate_name = defaultdict(list)   # maps rate name -> list of country codes
            current_rate = ""                  # the most recently seen non-blank rate name
            processed_rows = 0
            skipped_empty_country = 0

            for row in rows[1:]:   # rows[1:] skips the header row
                # Safely read the rate name and country code cells (guard against short rows)
                rn = row[rate_name_col] if rate_name_col < len(row) else None
                country = row[country_col] if country_col < len(row) else None

                # Update current_rate if this row has a non-blank rate name
                if rn is not None and str(rn).strip():
                    current_rate = str(rn).strip()

                # Skip rows with no country code (e.g. blank rows between sections)
                if country is None or (isinstance(country, str) and not str(country).strip()):
                    skipped_empty_country += 1
                    continue

                # Add this country code to the list for the current rate name
                country_str = str(country).strip()
                if country_str:
                    by_rate_name[current_rate].append(country_str)
                    processed_rows += 1

            print(f"[*] TXT Debug: processed country rows={processed_rows}")
            print(f"[*] TXT Debug: skipped rows with empty Country Code={skipped_empty_country}")
            print(f"[*] TXT Debug: distinct RateName groups={len(by_rate_name)}")

            # -----------------------------------------------------------------------
            # Build the output lines (RateName  code1, code2, ...).
            # Rate names are sorted alphabetically, with any blank rate name placed last.
            # -----------------------------------------------------------------------
            for rate_name in sorted(by_rate_name.keys(), key=lambda x: (x == "", x)):
                countries = by_rate_name[rate_name]
                display_rate = _normalize_zone_label(rate_name)
                line = f"{display_rate}  {', '.join(countries)}"
                lines.append(line)

            print(f"[*] TXT Debug: CountryZoning output lines={len(lines)}")
            if lines:
                print(f"[*] TXT Debug: first line preview={lines[0][:200]}")
            else:
                print("[WARN] TXT Debug: no CountryZoning lines generated from Excel")

    # -----------------------------------------------------------------------
    # Optional: append Demand Surcharge zone → country codes from extracted JSON
    # -----------------------------------------------------------------------
    if extracted_json_path:
        jp = Path(extracted_json_path)
        if jp.exists():
            try:
                with open(jp, encoding="utf-8") as jf:
                    data = json.load(jf)
                name_to_code = _load_country_codes()
                dsc = data.get("DemandSurchargeCountries") or []
                if dsc:
                    before = len(lines)
                    _append_demand_surcharge_countries_lines(lines, dsc, name_to_code)
                    print(
                        f"[*] TXT Debug: appended DemandSurchargeCountries "
                        f"({len(lines) - before} lines incl. separator)"
                    )
                ggc = data.get("GoGreenPlusCost") or []
                if ggc:
                    before = len(lines)
                    _append_gogreen_block_lines(lines, ggc, name_to_code)
                    print(
                        f"[*] TXT Debug: appended GoGreenPlusCost blocks "
                        f"({len(lines) - before} lines incl. separator)"
                    )
            except Exception as e:
                print(f"[WARN] TXT Debug: could not append DemandSurchargeCountries / GoGreen: {e}")
        else:
            print(f"[WARN] TXT Debug: extracted_json_path not found: {jp}")

    # Write all lines to the TXT file, separated by newlines
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text("\n".join(lines), encoding="utf-8")
    print(f"[OK] TXT Debug: wrote file {output_path}")
    return str(output_path)


def main():
    """
    Entry point when this script is run directly from the command line.

    Looks for the Excel file at output/DHL_Rate_Cards.xlsx (relative to the
    script's own folder) and saves the TXT file in the same output/ folder.
    """
    # Resolve paths relative to the folder where this script lives,
    # so the script works regardless of where it is run from
    script_dir = Path(__file__).resolve().parent
    excel_path = script_dir / "output" / "DHL_Rate_Cards.xlsx"
    output_path = script_dir / "output" / "CountryZoning_by_RateName.txt"

    print("Creating CountryZoning TXT from DHL_Rate_Cards.xlsx...")
    extracted_json = script_dir / "processing" / "extracted_data.json"
    out = create_country_region_txt(
        excel_path=str(excel_path),
        output_path=str(output_path),
        extracted_json_path=str(extracted_json) if extracted_json.exists() else None,
    )
    print(f"Saved: {out}")


# Only run main() when this script is executed directly (e.g. python country_region_txt_creation.py).
# Does NOT run when imported as a module by pipeline_main.py.
if __name__ == "__main__":
    main()
